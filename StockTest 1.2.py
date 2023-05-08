#!/usr/bin/env python
# coding: utf-8

# In[1]:


import yfinance as yf
import pandas as pd
import yahoo_fin.stock_info as si

import requests  
import json  

import numpy as np  
import matplotlib.pyplot as plt  
from scipy.signal import argrelextrema  
from datetime import datetime, timedelta  
from os import system 


# In[3]:


stock_list = pd.read_csv('/Users/Nash/Desktop/Trading/nasdaq_screener_1608593683363.csv')  
count = 0  
screen_list = []  
  
#Get the API Key that is stored in file  
file_key = open('/Users/Nash/Desktop/Trading/financialmodelingprep_api_key.txt','r')  
key = file_key.read()  
file_key.close()  
  
#Used if Stock runs into an exception (such as no revenue detected)  
def removeSym(sym):  
    global screen_list_df  
      
    mapping = screen_list_df['symbol'] != sym  
    screen_list_df = screen_list_df[mapping]  

#Gets the initial stocklist in form of JSON converted to DataFrame based on sector, exchange, and no dividend
def screenSector(sector, exchange, key):  
      
    r = requests.get('https://financialmodelingprep.com/api/v3/stock-screener?'  
                     'marketCapMoreThan=250000000&volumeMoreThan=800000&'  
                     'sector={}&exchange={}&dividendLowerThan=0&apikey={}'.format(sector,exchange,key))  
    r = r.json()  
    initScreen = pd.DataFrame(r)  
    return initScreen  

#Removes any stocks that do not meet the Volume Weighted Average Price threshold
def screenVWAP(stock_df):  
    mapping1 = stock_df['price']*stock_df['volume'] > 12000000  
    new_screen_list_df = stock_df[mapping1]  
      
    return new_screen_list_df  

#Simple method to remove any shell companies from screen DataFrame
def removeShellComp(stock_df):  
    mapping1 = stock_df['industry'] != 'Shell Companies'  
    new_screen_list_df = stock_df[mapping1]  
      
    return new_screen_list_df  

"""Initial Screening to get stocks that meet requirements for what would 
have taken hours (now only takes seconds)""" 

screen_list_df = pd.DataFrame()  
  
for sector in getTop5().index.values.tolist():  
    for exchange in ['NYSE','NASDAQ']:  
        screen_list_df = screenSector(sector, exchange, key).append(screen_list_df, ignore_index = True)  
        
        
screen_list_df = screenVWAP(screen_list_df)  
screen_list_df = removeShellComp(screen_list_df)  
new_screen_list_df = screen_list_df['symbol']  
# screen_list_df.set_index('symbol',inplace=True)  
screen_list_df


# In[5]:


#Simple method that gets the
def getRevenueGrowth(stock,key):  
    global screen_list_df  
      
    sym = stock  
    r = requests.get('https://financialmodelingprep.com/api/v3/income-statement/{}?limit=10&apikey={}'.format(stock,key))  
    r = r.json()
    print(type(r))
    print(r)  
#     try:  
    stock = pd.DataFrame.from_dict(r)  
    stock =  stock[['date', 'revenue', 'costOfRevenue','grossProfit',  
                    'costAndExpenses','ebitda','operatingIncome','netIncome']]  
    stock = stock.transpose()  

    rev_grwth1 = float(stock[0]['revenue']/stock[1]['revenue'] - 1)  
    rev_grwth2 = float(stock[1]['revenue']/stock[2]['revenue'] - 1)  
    print("RG1:",rev_grwth1)
    print("RG2:",rev_grwth2)

    opCost_grwth1 = float(stock[0]['costAndExpenses']/stock[1]['costAndExpenses'] - 1)  
    opCost_grwth2 = float(stock[1]['costAndExpenses']/stock[2]['costAndExpenses'] - 1)  
    print("OCG1:",opCost_grwth1)
    print("OCG2:",opCost_grwth2)

    rev_grwth = {sym:{'Revenue Growth':[rev_grwth1,rev_grwth2],'Cost Growth':[opCost_grwth1,opCost_grwth2]}}
    print(rev_gwth)
#     except:
#     print(sym)  
#     removeSym(sym)  
#         return  
      
      
    return rev_grwth  
  
s = {}  
s = s.update(new_screen_list_df.apply(lambda x: getRevenueGrowth(x,key)))  
getRevenueGrowth('GHIV',key)


# In[17]:


print(s)


# In[30]:


TODAY = datetime.now()
START = TODAY - timedelta(days=120)
print(TODAY)
print(START)
def listToString(list):  
    return " ".join(list)  

##NEEDS TO REMOVE DELISTED STOCKS
def removeDelisted(stock):
    pass
#See the Cell Block Below to Understand Logic; Gets Top Value Open/Close depending if day makes Gains/Loss  
def getGreaterArray(Open, Close):  
      
#     print(\"Open greater: \  np.greater(Open,Close))        # For Debbugging purposes  
#     print(\"Close greater: \  np.greater_equal(Close,Open)) # For Debbugging purposes  
    mask1 = np.greater(Open,Close)  
    mask2 = np.greater_equal(Close,Open)  
  
    Open_Greater = Open*mask1  
    Close_Greater = Close*mask2  
  
#     print('\n')  
#     print(Open_Greater)  
#     print(Close_Greater)  
  
    arr3 = Open_Greater+Close_Greater  
#     print(arr3)  
      
    return arr3  
  
#See the Cell Block Below to Understand Logic; Gets Lesser Value Open/Close depending if day makes Gains/Loss  
def getLesserArray(Open, Close):  
      
#     print(\"Open Lesser: \  np.less(Open,Close))        # For Debbugging purposes  
#     print(\"Close Lesser: \  np.less_equal(Close, Open)) # For Debbugging purposes  
    mask1 = np.less(Open,Close)  
    mask2 = np.less_equal(Close,Open)  
  
    Open_Less = Open*mask1  
    Close_Less = Close*mask2  
  
#     print('\n')  
#     print(Open_Less)   # For Debbugging purposes  
#     print(Close_Less)  # For Debbugging purposes  
 
    arr3 = Open_Less+Close_Less  
#     print(arr3)  # For Debbugging purposes  
      
    return arr3  
  
#Counts the streak of a stock's historical price to be able to determine gain/loss streak distribution  
#RETURNS a list of the gain/loss distribution  
def countStreak(series):  
      
    count = 0  
    to_store = 0  
    index = 0  
    streak = []  
    print(len(series))  
      
    for i in range(len(series)):  
          
        if (count == 0) & (series[i] == True):  
            count += 1  
        elif (series[i-1] == True) & (series[i] == True):  
            count += 1  
        else:  
            to_store = count  
            count = 0  
              
        if count == 0:  
            streak.append(to_store)  
            if to_store != 0:  
                streak.append(count)  
            index += 1  
          
        if (i == len(series) - 1) & (series[i] == True):  
            streak.append(count)  
          
    return streak  
  
# Gets summary statistics of an array or list passed in in terms of wins and losses; returns dictionary  
def countGainLoss(series):  
    count = 0  
    for i in series:  
        if i > 0:  
            count+=1  
      
    stats = {'Gains': count, 'Losses': len(series) - count, 'Total': len(series), 'WinLoss': count/(len(series) - count)}  
     
    return stats  
  
# Takes dataframe of stock prices (DF is 'stock') and uses minimum or maximum (max is the default)  
def identifyLevel(stock, col = 'Pivot', name = "Default Stock"):  
    print("IDENTIFY LEVEL\n")    
    level_df = pd.DataFrame()  
    ind = 1       #Number that identifies different column name; changes on each iteration of loop  
    count = 0     #Stops the while loop when all the prices have been checked for
    mask = ~stock['Pivot'].isna()
    stock = stock[mask]
    remaining_df = stock
    
    print("FROM IDENTIFY, ", stock)  
    
    #Each iteration determined by the creation of a new column 
    #(index is date, so each level's dates are unique from last 
    # -> increases count on each iteration)
    #vvvvvvvvvvvvvvvvvvvvvvvvvvv
    while count < stock.shape[0]: #Replace series_max with the dataframe the function takes in; shape gives length  
      
#       print('iteration: ', ind)  
          
        price = remaining_df[col].iloc[0] #Grabs first price in 'Remaining_df' to use for Range Mask
        ##FIXED BUG: Originally 'stock[col]'; This would put number in two levels and create duplicate indexes in finalizeStock
        mask_x = (price*0.9875 <= remaining_df[col]) & (remaining_df[col] <= price*1.0125) #Creates a mask of prices in the DF within range  
#         print(mask_x)  
        temp_df = pd.DataFrame({'Level {}'.format(ind) : remaining_df[col][mask_x].values.tolist()}  #Takes the Level Line for values that match current 'price' being searched for  
                               ,index = remaining_df[col][mask_x].index)  
        level_df = pd.concat([level_df,temp_df],axis=1) #Applies Mask to a level series creating a new column (New DF)  
  
        remaining_df = remaining_df[~mask_x]  #Removes the values that were just applied at this level from remaining_df
  
        ind += 1      #Iterates to change column name for next go around  
        count = level_df.shape[0]  
#     print(level_df)
          
            #print('An error occurred in \"identifyLevel\" function.  Check to make sure {} has \"min\" or \"max\" column.'.format(name))  
            #return  
#     print("Level Shape", df.shape)      
    return level_df  
  
#Takes in the stock_df ('df') with the minima and maxima points and the 'level_df'  
#Replaces the minima/maxima points with the average value if the level contains 2 or more points  
#Creates a Column of the level's strength:  
    #(Pivot - 1 point and less than 30 days, Weak - 2 or more Points & More than 30 Days,   
    # Relevant - 2 Points & Less than 30 Days, Strong - 3 Points and At least one Less than 30 Days,   
    # Very Strong - 4 or More Points and At least one less than 30 Days)  
def finalizeStock(stock, pivot_df, level_df):  
#     Going to need to check how many non-NaN values in each column for Level_DF  
#     If non-NaN value > 1, Average the values replace the values in df that give that average  
#     Use the TimeDelta data in Level_df Index as well as number of points to create the relevant label  
#         Create a Column called 'Average' in Level_Df and put average values there  
#         Repeat step above but call column 'Strength' and place the labels there  
#         Use 'Average' values and Replace  Maxima/Minima values in df that correspond to index  
#         Repeat step above merging the 'Strength' column above with corresponding index  
    print("FINALIZE STOCK\n")
    
    today = datetime.now()      
    num_of_pivots = level_df.shape[0] - level_df.isna().sum()  # Stores the number of pivots for each level line  
    print(num_of_pivots)  
  
    temp = pd.DataFrame()  
    
    #4/19/21: NEED TO IMPLEMENT TO REMOVE DELISTED STOCKS
    try:
        for i,piv in num_of_pivots.iteritems():   # Lazily loops through num_of_pivots (i is the name of the Level i.e. 'Level 1', piv is the numb of pivs)  

            pivot_dates = pd.DataFrame(level_df[~level_df.isna()[i]].index)  # Creates a DF of the Pivot Points Dates (Col: 'Date')  
            pivot_days = ((today - pivot_dates['Date'])/np.timedelta64(1, 'D')).astype(int)  # Calculates days between pivot date and today  
            pivot_dates['Days'] = pivot_days  
            pivot_dates.set_index('Date',inplace = True)   
            check_date = (pivot_dates <= 30).any().values[0]  # If any values <= 30, then TRUE  

            avg = level_df[i].mean(axis = 0, skipna = True)  # If there is more than 1 pivot, get the average  

            if piv > 1:  

                if (piv == 2) & check_date:  
                    listLabel = ['Relevant']*piv  

                elif (piv >= 2) & ~check_date:  
                    listLabel = ['Weak']*piv  

                elif (piv == 3) & check_date:  
                    listLabel = ['Strong']*piv  

                elif (piv >= 4) & check_date:  
                    listLabel = ['Very Strong']*piv  

            elif (piv == 1) & check_date:  
                listLabel = ['Pivot']*piv  

            else:  
                listLabel = ['Old Pivot']*piv  

            pivot_dates['Level Type'] = listLabel  
            pivot_dates['Level'] = avg  

            temp = pd.concat([temp, pivot_dates])
    #     print("DF", df)
    #     print("TEMP", temp)
        merged = pd.merge(df, temp, on='Date',how = 'outer') #CHANGED: Was 'outer' #UPDATE: now 'outer' again?
    #     print("MERGED 1: ", merged)
        final_df = pd.merge(stock, merged[['Days', 'Level Type', 'Level']], on='Date', how='outer') #CHANGED: Was 'outer' #UPDATE: now 'outer' again?
    except:
        removeDelisted(stock)
        return
        
#     print("DF", df)
#     print("FINALIZED: ", final_df)  
#     print("Finalized Shape", final_df.shape)
    
    #check_watchlist(final_df)
    
    return final_df  

#Gets the minima and maxima values, as well as combining them into a 'Pivot' column
def grabPivots(df,period = 5):  
    print("GRAB PIVOTS\n")  
    #Tried simply merging a DataFrame of Min and Max together as an inner, but that did not work  
    df['min'] = df.iloc[argrelextrema(getLesserArray(df.Open.values,df.Close.values), np.less_equal,  
                    order=period)[0]]['Close'] #Gets local minima points for given period on 'Close' prices  
    df1 = pd.DataFrame(df['min']) 
    df1.rename(columns={'min':'Pivot'},inplace = True)  
    mask = ~df['min'].isna()
      
    df1 = df1[mask]  
    print("DF1 MIN:",  df1)  
      
    df['max'] = df.iloc[argrelextrema(getGreaterArray(df.Open.values,df.Close.values), np.greater_equal,  
                    order=period)[0]]['Close'] #Gets local maxima points for given period on 'Close' prices  
    df2 = pd.DataFrame(df['max'])
    df2.rename(columns={'max':'Pivot'},inplace = True)
    ##FIXED BUG: AMCI had error on 10/28 with local minima being the local maxima as well duplicating rows
    mask2 = (~df['max'].isna()) & (df['min'] != df['max']) 
      
    df2 = df2[mask2]  
    print("DF2 Max: ",  df2)  
  
    df3 = pd.concat([df1, df2])  
    print("DF3 Concat: ",  df3)  
    df = df.merge(df3, on='Date', how = 'outer',sort = True) #Basically takes the min and max values, puts into one column called 'Pivot'  
                                                 #This will be used for checking levels in identifyLevels()  
    print("GP Shape", df.shape)
    return df  
  
#Example of Grabbing large amount of stocks as thread and reading price  
df1 = pd.read_excel("/Users/Nash/Desktop/Trading/Master_Stock_Screen.xlsx",  sheet_name='2020-12-24')  
df1= df1['Symbol'][:100]  
list = df1.values.tolist()  
print(list)  
data = yf.download(listToString(list), start=START,   
                    end=TODAY,  group_by='tickers')  
  
  
  
# This Works for a single Stock.  Now I need to see if I can get it to work for an entire DataFrame!!!

############
# df = stock_test['A']  
# # df = pd.DataFrame(df)  

  
# df_pivots = grabPivots(df,period = 5)  #Get the minimum and maximum values of stockprice
#                                  # Add the ability to change time period and time frame of stock

# # df_pivots = df_pivots[df_pivots['Pivot'] > 0] 
# level_df = identifyLevel(df_pivots, name = 'A')  

# df = finalizeStock(df,df_pivots, level_df)  
############

# plt.scatter(df.index, df['min'], c='r')  
# plt.scatter(df.index, df['max'], c='g')  
# plt.plot(df.index, df['Close'])  
# plt.show() 

# map(lambda i: grabPivots(i),np.unique(stock_test.columns.get_level_values(0).tolist()))  
#stock_test.apply(lambda x: x[map(lambda i: grabPivots(i),np.unique(stock_test.columns.get_level_values(0).tolist()))])  


# In[31]:


from timeit import default_timer as timer
start = timer()

streak = []  
count = 0  
  
cumlative = pd.DataFrame()

for i in np.unique(stock_test.columns.get_level_values(0).tolist()):  
      
    count += 1  
    print("STOCK {}".format(i))  
    df = stock_test[i]
#     print("1.", df.shape)
    df_pivots = grabPivots(df, period = 5) #ABB ISSUE, Adds 2 Rows to df
    level_df = identifyLevel(df_pivots, name = i)  
    
    df = finalizeStock(df,df_pivots, level_df) #ABB ISSUE, Adds 2 Rows to df
#     print("2.", df.shape)
    df.columns = pd.MultiIndex.from_product([[i],df.columns])
#     print("3.", df.shape)
 #Creates a new column called 'ticker' that takes an array the size of all the rows for the stock.  
    #df.set_index([df.index, md],inplace = True)  
    if count == 1:  
        cumlative = df  
    else:
#         print(i,"'s SHAPE'")
#         print("4. DF SHAPE: ", df.shape)
#         print("CUM SHAPE: ", cumlative.shape)
        
        cumlative = pd.concat([cumlative,df],axis=1)  
print("df of {}: ".format(i), cumlative)  
      
#     plt.scatter(df.index, df['min'], c='r')  
#     plt.scatter(df.index, df['max'], c='g')  
#     plt.plot(df.index, df['Open'])  
# plt.show()  
#     GET HISTORICAL GAIN STREAK  
#     win_streak = countStreak(np.greater_equal(stock_test[i].Open.values,stock_test[i].Close.values))  
#     loss_streak = countStreak(np.less_equal(stock_test[i].Open.values,stock_test[i].Close.values))  
#     print(\"------- {} -------\".format(i))  
#     print(\"MAX GAIN STREAK: \  max(win_streak))  
#     print(\"MAX LOSS STREAK: \  max(loss_streak))  
#     plt.hist(win_streak, density = True, bins=max(win_streak),align = 'mid')  
#     plt.ylabel('Probability')  
#     plt.xlabel('Daily Gain Streak')  
  
#     stock_test.apply(lambda x: grabPivots(x[i]))

end = timer()
print(end - start, " seconds") # Time in seconds, e.g. 5.38091952400282


# In[61]:


from sklearn.linear_model import LinearRegression

watchlist = []

def get_currentPrice(stock, ticker):
    return pd.DataFrame(stock.iloc[-1]).T

def screen_watchlist(stock, ticker):
    
    global watchlist
    
    current_price = stock.iloc[-1][3]  #Gets the Close price from last value of df DataFrame to initialize current_price
    mask = (stock['Level Type'] == 'Pivot') | (stock['Level Type'] == 'Relevant') | (stock['Level Type'] == 'Strong') | (stock['Level Type'] == 'Very Strong')
    screen_df = stock[mask]
    print(screen_df)

def get_trendline(stock, ticker):
    
    current_price = get_currentPrice(stock, ticker)

    minimum_df = pd.DataFrame(stock['min'])
    min_mask = ~minimum_df['min'].isna()
    minimum_df = minimum_df[min_mask]

    maximum_df = pd.DataFrame(stock['max'])
    max_mask = ~maximum_df['max'].isna()
    maximum_df = maximum_df[max_mask]
    
    if not (minimum_df['min'].empty):
        min_date = minimum_df.index[-1] # Gets minimum_df's last date
    else:
        min_date = datetime.min
    if not (maximum_df['max'].empty):
        max_date = maximum_df.index[-1] # Gets maximum_df's last date
    else:
        max_date = datetime.min

    #Checks if min or max is the last extrema, if no min or max, return "No Trend"
    if(min_date > max_date):
        last_extreme = min_date
    elif (min_date < max_date):
        last_extreme = max_date
    else:
        return ["No Trend"]

    print(minimum_df)
    print(maximum_df)

    #Checks if the last extrema is the same as the current price (This indicates only one point and therefore no trend)
    if(stock.loc[last_extreme].name != current_price['Close'].index[0]):

        #Checks to see if current price is already in the "stock" DataFrame
        if(stock.loc[current_price.index[0]].empty):
            trend = pd.concat([stock.loc[last_extreme:],current_price])
        else:
            trend = stock.loc[last_extreme:]

        X = (trend.index -  trend.index[0]).days.values.reshape(-1, 1)
        Y = trend['Close'].values.reshape(-1,1)

        print(X)
        linear_regressor = LinearRegression()  # create object for the class
        linear_regressor.fit(X, Y)  # perform linear regression
        Y_pred = linear_regressor.predict(X)  # make predictions

        print("SCORE: ",linear_regressor.score(X,Y))
        print("COEF_: ",linear_regressor.coef_)

        plt.scatter(X, Y)
        plt.plot(X, Y_pred, color='red')
        plt.show()

        return ["Trend", linear_regressor.fit(X, Y), X, pd.DataFrame(trend['Close'])]

    else:
        return["No Trend"]
    
    
    
def get_resistance(stock, ticker):
    
    current_price = get_currentPrice(stock, ticker)
    
    #RESISTANCE: Checks for level type, prioritizing 'Very Strong', and if any are greater than "current_price"
    for level_type in ['Very Strong', 'Strong', 'Relevant', 'Pivot']:
        resistance_mask = ((current_price['Close'][0] < stock[stock['Level Type'] == level_type]['Level']) 
                           & (stock['Level Type'] == level_type))

        if(resistance_mask.any()):
            print("\n",level_type,"\n")

            resistance_lvl = (stock[resistance_mask]['Level'].min(), level_type) #Sets the resistance line (As Tuple) to watch as the level currently closest to current_price

            print(level_type, "Resistance")
            print(resistance_lvl)
            
            return resistance_lvl
        
def get_support(stock, ticker):
    
    current_price = get_currentPrice(stock, ticker)
    
    #SUPPORT: Checks for level type, prioritizing 'Very Strong', and if any are less than "current_price"
    for level_type in ['Very Strong', 'Strong', 'Relevant', 'Pivot']:
        support_mask = ((current_price['Close'][0] > stock[stock['Level Type'] == level_type]['Level']) 
                           & (stock['Level Type'] == level_type))

        if(support_mask.any()):
            print("\n",level_type,"\n")

            support_lvl = (screen_df[support_mask]['Level'].max(), level_type) #Sets the resistance line (As Tuple) to watch as the level currently closest to current_price
            print(level_type, "Support")
            print(support_lvl)
            
            return support_lvl

n = 'AAL'
stock = cumlative[n]
get_trendline(stock,n)

trend_df = pd.DataFrame()

for n in np.unique(stock_test.columns.get_level_values(0).tolist()):
    stock = cumlative[n]
    
    print(n)
    
    trend = get_trendline(stock,n)
    
    temp_df = pd.DataFrame(trend).T
    if trend[0] != 'No Trend':
        trend_df = pd.concat([trend_df,temp_df])


# In[71]:


#Initial Trend Length
sum = 0

for i in trend_df[2].values.tolist():
    sum += len(i)
length = len(trend_df[2].values.tolist())
print(sum/length)


# In[37]:


cumlative['ABCL']


# In[ ]:


streak = countStreak(np.greater_equal(stock_test['ABR'].Open.values,stock_test['ABR'].Close.values))  
print("ABR : ",  streak)  
stats = countGainLoss(np.greater_equal(stock_test['ABR'].Open.values,stock_test['ABR'].Close.values))  
print('Win Percentage: ', stats.get("Gains")/stats.get("Total"))  
print('Win/Loss: ', stats.get("WinLoss"))  
print("MAX: ",  max(streak))  
plt.hist(streak, density=True, bins=max(streak))  
print("")  
  
np.unique(stock_test.columns.get_level_values(0).tolist())  
  
# print(np.unique(stock_test.columns.get_level_values(0).tolist()))  
# reg = stock_test.loc[:,stock_test.columns.get_level_values(1).isin({'Adj Close'})]  
# high = reg  
# if reg > high:  
#     high = reg  
# else:  
#     high = reg.loc[1:,stock_test.columns.get_level_values(1).isin({'Adj Close'})]  
# high


# In[26]:


stock_test = data
stock_test


# In[ ]:


## SIMULATE EXPECTED DISTRIBUTION OF WIN STREAKS (TRUE) GIVEN PROBABILITY  
streak = []  
probability = 0.546  
  
for i in range(10001):  
    x = np.random.random()  
    if x < probability:  
        streak.append(True)  
    else:  
        streak.append(False)  
print(streak)  
t = countStreak(streak)  
print(t)  
print(max(t))  
plt.hist(t, density = True, bins=max(t),align = 'mid')  
plt.ylabel('Probability')  
plt.xlabel('Daily Gain Streak')


# In[ ]:


print(stock_test.columns)  
print(cumlative.columns)


# In[4]:


import requests  
import xlrd  
from openpyxl import Workbook as WB  
  
def getOnlineFile(url, filename):  
  
    r = requests.get(url, allow_redirects=True)  
    with open(filename, 'wb') as f:  
        f.write(r.content)  
#Need to Fix (doing this to make it easier for getting Top 5 Shiller)  
def flipDataFrame(df):  
      
    index_values = df.columns.values.tolist()  
    col_values = df.index.values.tolist()  
      
    temp_pd = pd.DataFrame(columns = col_values)  
    temp_pd.set_index(index_values, inplace = True)  
    print(temp_pd)  

#Simple "Getter Method"
def getShillerRelative():  
    global shiller_rel_df  
    return shiller_rel_df  
  
#Initializes the ShillerRelative Value  
def setShillerRelative(filename):  
    global shiller_rel_df  
      
    shiller_df = pd.read_excel(filename, engine = 'xlrd')  
      
    ##Cleaning the data by dropping first row and regular PE ratios columns  
    shiller_df.drop(0, inplace = True)  
    shiller_df.rename(columns={'Unnamed: 0': 'Date'}, inplace = True)  
    
    ##Removes even columns it appears, not sure why.  Perhaps because they're empty?
    for i in range(2,24,2):  
        shiller_df.drop('Unnamed: {}'.format(i), axis = 1, inplace = True)  
  
    ##Indexing the data to be greater than this date for daily.  
    mask = shiller_df['Date'] >= '2013-07-12'  
    #print(shiller_df.loc[mask])  
  
    shiller_df = shiller_df.loc[mask]  
    shiller_df.set_index('Date', inplace = True)

    shiller_rel_df = shiller_df[0:1]/shiller_df.mean()  

#Returns the Top 5 Shiller Relative PE values from the sector as a dataframe and prints them to the screen.  
def getTop5():  
    global shiller_rel_df  
      
    column_name = getShillerRelative().index.tolist()[0]  
      
    ##Transpose the Shiller Relative Dataframe so Sectors are the Indices  
    top5_df = shiller_rel_df.transpose()  
    top5_df.rename(columns={column_name:'Shiller Relative PE'},inplace =True)  
    top5_df.columns.name = 'Sector'  
#     print(top5_df)  
    ##Create list form of the Shiller DataFrame to filter values  
    top_5 = []  
      
    _min = top5_df.values.tolist()[0][0]  
    new_shiller = []  
    shiller_list = top5_df.values.tolist()  
      
    #Flatten the list (Not necessary, but okay)  
    for sublist in shiller_list:  
        for value in sublist:  
            new_shiller.append(value)  
      
    #Gets the best 5 relative values  
    for i in range(5):  
          
        for j in new_shiller:  
              
            if j < _min:  
                _min = j  
        top_5.append(_min)  
        new_shiller.remove(_min)  
        _min = 1000000000000000  
      
    ##Filters the DataFrame to only the top 5 best relative values  
    top5_df = top5_df[top5_df['Shiller Relative PE'].isin(top_5)]  
      
    return top5_df  
     
def shillerRelative_toExcel(filename):  
    shillerWB = load_workbook(filename)  
    ws_data = ws.active  
    ws_data.title = 'Sector Data'  
      
      
    shillerWB.create_sheet('Relative Shiller P/E')  
  
shiller_rel_df = pd.DataFrame()  
  
url = 'https://www.gurufocus.com/download_sector_shiller_pe.php'  
filename = '/Users/Nash/Desktop/Trading/sector_shiller_pe.xls'  
  
#print(\"initial:\ \ \  getShillerRelative())  
#getOnlineFile(url, filename)  
setShillerRelative(filename)  
#print(\"after:\ \ \  getShillerRelative(),\"\ \")  
getTop5().index.values.tolist() 

screen_list_df = pd.DataFrame()  
  
#Initial Screening to get stocks that meet requirements for what would have taken hours (now only takes seconds)  
for sector in getTop5().index.values.tolist():
    for exchange in ['NYSE','NASDAQ']:
        screen_list_df = screenSector(sector, exchange, key).append(screen_list_df, ignore_index = True)  
 
screen_list_df

# ## Mapping example if I wanted to create seperate dataframes on Sectors  
# mapping = screen_list_df['sector'] == 'Industrials'  
# sector1 = screen_list_df[mapping]  
# sector1  
  
mapping2 = screen_list_df['price']*screen_list_df['volume'] > 12000000  
new_screen_list_df = screen_list_df[mapping2]  
new_screen_list_df
  

from datetime import date  
from openpyxl import Workbook, load_workbook  
  
path = r'/Users/Nash/Desktop/Trading/Master_Stock_Screen.xlsx'  
  
wb = Workbook()  
ws = wb.active  
  
ws['A1'] = 'This is the Master File for screened Stocks'  
ws['A2'] = 'Program last run: ' + str(date.today())  
  
wb.save(path)  
  
df = pd.DataFrame(screen_list)  
  
book = load_workbook(path)  
writer = pd.ExcelWriter(path, engine = 'openpyxl')  
writer.book = book  
  
if str(date.today()) not in book.sheetnames:  
    df.to_excel(writer, sheet_name = str(date.today()))  
writer.save()  
writer.close()


# In[ ]:




